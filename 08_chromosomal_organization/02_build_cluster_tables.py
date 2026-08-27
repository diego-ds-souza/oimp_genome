#!/usr/bin/env python3
"""
Full membership tables for the tandem clusters and multi-family superloci
(Tables S16-S18), with per-gene differential expression joined.

Step 01 builds its family table by reading one gene list per family and joining
them onto the gene positions, so a gene listed under two families produces two
rows and is clustered twice, once under each. The 16 families cover 1,250 rows
across 1,248 distinct genes: two genes are listed as both GR and OR.

The published tables collapse each gene to a single primary family first - OR
takes precedence over GR, and UGT and SerPro over CAZyme - which gives 175
clusters, 865 clustered genes and 7 superloci. This script reproduces that
collapse and then re-derives the clusters and superloci with the same rules as
step 01, so its output matches the published tables. It ends by checking its
own cluster counts against Table S17 and reporting any mismatch.

Rules, as in step 01:
    cluster     within one family and chromosome, genes are chained while
                chr_start[i] - chr_end[i-1] <= MAX_CLUSTER_DISTANCE, and at
                least MIN_CLUSTER_GENES are required
    superlocus  clusters on a chromosome are walked in coordinate order and
                merged while the next starts within MAX_SUPERLOCUS_GAP of the
                running end; a merged run counts as a superlocus if it spans
                more than one family

Run from the repository root:
    conda activate oimp_08_chromosomal_organization
    python 08_chromosomal_organization/02_build_cluster_tables.py
"""

import csv, os
from collections import defaultdict, OrderedDict, Counter
import openpyxl

# --------------------------------- settings ----------------------------------
BASE = os.environ.get("BASE", "results/08_chromosomal_organization")
OUT = os.environ.get("OUT", "results/08_chromosomal_organization/tables")
SUPP4 = os.environ.get("SUPP4", "data/Supplementary_file_4_DE_Target_families.xlsx")
S17 = os.environ.get("S17", "data/Supplementary_Tables_S1-S18.xlsx")

MAX_CLUSTER_DISTANCE = 100_000   # bp between consecutive genes of one cluster
MIN_CLUSTER_GENES = 2            # genes required to call a cluster
MAX_SUPERLOCUS_GAP = 50_000      # bp between clusters of a superlocus
# -----------------------------------------------------------------------------
# when a gene is listed under two families, keep the one with the lower rank
# (CAZyme yields to UGT and SerPro; GR yields to OR - this reproduces Table S17)
FAMILY_RANK = defaultdict(lambda: 0, {'CAZyme': 2, 'GR': 1})

GROUP_OF = {}

# 1) Read the chromosome-level gene positions from step 01.
with open(f'{BASE}/gene_positions_chromosome_level.tsv') as fh:
    positions = list(csv.DictReader(fh, delimiter='\t'))

rows_by_gene = defaultdict(list)
for r in positions:
    if r['family'] not in ('NA', ''):
        rows_by_gene[r['gene_id']].append(r)
        GROUP_OF[r['family']] = r['group']

primary, all_families = {}, defaultdict(set)
for g, rs in rows_by_gene.items():
    for r in rs:
        all_families[g].add(r['family'])
    primary[g] = min(rs, key=lambda r: FAMILY_RANK[r['family']])

# 2) Collapse each gene to one primary family, then chain genes into clusters.
by_fam_chr = defaultdict(list)
for g, r in primary.items():
    if r['chromosome'] not in ('NA', ''):
        by_fam_chr[(r['family'], r['chromosome'])].append(r)

clusters = []
for (fam, chrom), genes in by_fam_chr.items():
    genes.sort(key=lambda r: int(r['chr_start']))
    run = [genes[0]]
    for prev, cur in zip(genes, genes[1:]):
        if int(cur['chr_start']) - int(prev['chr_end']) <= MAX_CLUSTER_DISTANCE:
            run.append(cur)
        else:
            if len(run) >= MIN_CLUSTER_GENES:
                clusters.append((fam, chrom, run))
            run = [cur]
    if len(run) >= MIN_CLUSTER_GENES:
        clusters.append((fam, chrom, run))

# stable ids: family group, then chromosome number, then start
CHRN = lambda c: int(c.split('_')[1])
clusters.sort(key=lambda t: (t[0], CHRN(t[1]), int(t[2][0]['chr_start'])))
CL = []
for i, (fam, chrom, genes) in enumerate(clusters, start=1):
    start = min(int(g['chr_start']) for g in genes)
    end   = max(int(g['chr_end'])   for g in genes)
    CL.append({'cluster_id': i, 'family': fam, 'group': GROUP_OF[fam],
               'chromosome': chrom, 'start': start, 'end': end,
               'n_genes': len(genes), 'span_kb': round((end - start) / 1000, 1),
               'genes': genes})

# 3) Merge nearby clusters, keeping runs that span more than one family.
SL = []
for chrom in sorted({c['chromosome'] for c in CL}, key=CHRN):
    chr_cl = sorted([c for c in CL if c['chromosome'] == chrom], key=lambda c: c['start'])
    i = 0
    while i < len(chr_cl):
        run, end = [chr_cl[i]], chr_cl[i]['end']
        j = i + 1
        while j < len(chr_cl) and chr_cl[j]['start'] <= end + MAX_SUPERLOCUS_GAP:
            if chr_cl[j]['family'] != run[0]['family']:
                run.append(chr_cl[j]); end = max(end, chr_cl[j]['end'])
            j += 1
        if len({c['family'] for c in run}) > 1:
            SL.append({'superlocus_id': len(SL) + 1, 'chromosome': chrom,
                       'start': min(c['start'] for c in run),
                       'end': max(c['end'] for c in run),
                       'n_clusters': len(run),
                       'families': ' + '.join(dict.fromkeys(c['family'] for c in run)),
                       'total_genes': sum(c['n_genes'] for c in run),
                       'span_kb': round((max(c['end'] for c in run)
                                         - min(c['start'] for c in run)) / 1000, 1),
                       'clusters': run})
        i = max(j, i + 1)

# 4) Join the differential expression calls onto each gene.
SHEETS = OrderedDict([
    ('Developmental (Larva vs Adult)', {None: 'dev'}),
    ('Tissue (Antenna vs Thorax)',     {None: 'tissue'}),
    ('Sex (Male vs Female)',           {'Adults pooled': 'sex',
                                        'Antennae only': 'sex_antennae',
                                        'Thorax only':   'sex_thorax'}),
])
TAGS = ['dev', 'tissue', 'sex', 'sex_antennae', 'sex_thorax']
DE_COLS = [f'{t}_{s}' for t in TAGS for s in ('log2FC', 'direction', 'padj')]

de = defaultdict(dict)
wb = openpyxl.load_workbook(SUPP4, read_only=True, data_only=True)
for sheet, tagmap in SHEETS.items():
    tag, idx = tagmap.get(None), None
    for row in wb[sheet].iter_rows(values_only=True):
        if row is None or all(v is None for v in row):
            idx = None; continue
        first = row[0]
        if isinstance(first, str) and first.strip() in tagmap:
            tag, idx = tagmap[first.strip()], None; continue
        # header row: identified by its content, because the first cell is
        # 'Rank' in some versions of the file and '1' in others
        if any(str(h).strip() == 'Gene ID' for h in row if h is not None):
            idx = {str(h).strip(): i for i, h in enumerate(row) if h is not None}
            continue
        if idx is None or row[idx['Gene ID']] is None:
            continue
        gid = str(row[idx['Gene ID']]).strip()
        de[gid][f'{tag}_log2FC']    = row[idx['log2FC']]
        de[gid][f'{tag}_direction'] = row[idx['Direction']]
        de[gid][f'{tag}_padj']      = row[idx['padj']]

# 5) Write the three tables.
def gene_block(g):
    d = de.get(g['gene_id'], {})
    b = OrderedDict([
        ('gene_id', g['gene_id']), ('gene_name', g['name']),
        ('scaffold', g['scaffold']), ('chr_start', g['chr_start']),
        ('chr_end', g['chr_end']), ('strand', g['chr_strand']),
        ('all_family_assignments', '+'.join(sorted(all_families[g['gene_id']]))),
        ('product', g['product']),
    ])
    for col in DE_COLS:
        b[col] = d.get(col, '')
    b['DE_any'] = 'yes' if d else 'no'
    return b

gene_rows = []
for c in CL:
    for k, g in enumerate(sorted(c['genes'], key=lambda g: int(g['chr_start'])), start=1):
        row = OrderedDict([('cluster_id', c['cluster_id']), ('family', c['family']),
                           ('group', c['group']), ('chromosome', c['chromosome']),
                           ('cluster_start', c['start']), ('cluster_end', c['end']),
                           ('cluster_n_genes', c['n_genes']),
                           ('cluster_span_kb', c['span_kb']),
                           ('position_in_cluster', k)])
        row.update(gene_block(g))
        gene_rows.append(row)

cluster_rows = []
for c in CL:
    d = [de.get(g['gene_id'], {}) for g in c['genes']]
    cnt = lambda t, v: sum(1 for x in d if x.get(f'{t}_direction') == v)
    cluster_rows.append(OrderedDict([
        ('cluster_id', c['cluster_id']), ('family', c['family']), ('group', c['group']),
        ('chromosome', c['chromosome']), ('start', c['start']), ('end', c['end']),
        ('n_genes', c['n_genes']), ('span_kb', c['span_kb']),
        ('n_DE_any', sum(1 for x in d if x)),
        ('n_adult_biased', cnt('dev', 'Adult')), ('n_larva_biased', cnt('dev', 'Larva')),
        ('n_antenna_biased', cnt('tissue', 'Antenna')), ('n_thorax_biased', cnt('tissue', 'Thorax')),
        ('n_male_biased', cnt('sex', 'Male')), ('n_female_biased', cnt('sex', 'Female')),
        ('genes', ','.join(g['gene_id'] for g in c['genes'])),
    ]))

sl_rows = []
for s in SL:
    for c in s['clusters']:
        for g in sorted(c['genes'], key=lambda g: int(g['chr_start'])):
            row = OrderedDict([('superlocus_id', s['superlocus_id']),
                               ('chromosome', s['chromosome']),
                               ('superlocus_start', s['start']), ('superlocus_end', s['end']),
                               ('superlocus_families', s['families']),
                               ('superlocus_total_genes', s['total_genes']),
                               ('superlocus_span_kb', s['span_kb']),
                               ('cluster_id', c['cluster_id']), ('cluster_family', c['family'])])
            row.update(gene_block(g))
            sl_rows.append(row)

os.makedirs(OUT, exist_ok=True)
def write(name, rows):
    with open(f'{OUT}/{name}', 'w', newline='') as fh:
        w = csv.DictWriter(fh, fieldnames=list(rows[0].keys()), delimiter='\t')
        w.writeheader(); w.writerows(rows)
    print(f'  {name}  ({len(rows)} rows)')

print('written:')
write('tandem_clusters_all_genes.tsv', gene_rows)
write('tandem_clusters_summary.tsv', cluster_rows)
write('superloci_all_genes.tsv', sl_rows)

# 6) Check the cluster counts against Table S17 and report any mismatch.
clustered = {g['gene_id'] for c in CL for g in c['genes']}
wb2 = openpyxl.load_workbook(S17, read_only=True)
s17 = Counter((r[1], r[3].replace(' ', '_'), int(r[6]))
              for r in wb2['Table S17'].iter_rows(min_row=3, values_only=True) if r[0] is not None)
mine = Counter((c['family'], c['chromosome'], c['n_genes']) for c in CL)
print(f"""
VALIDATION against the published supplementary tables
  clusters ................. {len(CL):>5}   Table S17: {sum(s17.values())}
  genes in clusters ........ {len(clustered):>5}   manuscript: 865
  family members ........... {len(primary):>5}   Table S16 total: 1250 (sums families; 2 genes are GR+OR)
  pct clustered ............ {100*len(clustered)/len(primary):>5.1f}%  manuscript: 69.3%
  superloci ................ {len(SL):>5}   Table S18: 7
  cluster mismatches vs S17: {list((s17 - mine).items()) + list((mine - s17).items())}
""")
